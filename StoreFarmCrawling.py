from openpyxl import Workbook
from bs4 import BeautifulSoup
import gui
from selenium import webdriver
from datetime import date
import os
from multiprocessing import Pool

# 헤드리스 모드 진행시 블록아웃을 삭제
checkList_info = ("아이디", "이름", "집주소", "전화번호", "통관번호")
customer_info_arr = []  # Customer 클래스의 배열

options=webdriver.ChromeOptions()
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
driver=webdriver.Chrome('C:\python/chromedriver',options=options)

class Customer:

    def __init__(self, id, name, address, phone, custom_num):
        self.id = id
        self.name = name
        self.address = address
        self.phone = phone
        self.custom_num = custom_num

    def get_customer_info(self):
        info = [0, self.id, self.name, self.address, self.phone, self.custom_num]
        return info

    def set_cutomer_info(self, id, name, address, phone, custom_num):
        self.id = id
        self.name = name
        self.address = address
        self.phone = phone
        self.custom_num = custom_num

def make_excel(url, customer_count):
    wb = Workbook()
    ws1 = wb.active
    # ws2= wb.create_sheet(title="second_sheet")  새로운 시트 파일 하나 더 추가하기
    for col in range(1, 6):
        ws1.cell(row=1, column=col, value=checkList_info[col - 1])
    # 이름, 전번, 집주소, 통관번호 입력을 넣어준다
    for row in range(2, customer_count + 2):
        customer_info = customer_info_arr[row - 2].get_customer_info()
        for col in range(1, 6):
            ws1.cell(row=row, column=col).value = customer_info[col]
    wb.save(url)
    wb.close()


def get_excel_file_name():
    return date.today().isoformat() + '.xlsx'


def add_customer(id, name, address, phone_num, custom_num):
    customer = Customer(id, name, address, phone_num, custom_num)
    customer_info_arr.append(customer)


def open_info_pages(url):
    try:
        driver.implicitly_wait(3)
        driver.get(url)
        read_spec_order()
    except Exception as ex:
        print("상세정보창을 열 수 없습니다",ex)
        exit()


def read_spec_order():
    id = driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[1]/tbody/tr[4]/td').text
    name = driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[1]/td').text
    phone_num = driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[2]/td[1]').text
    address = driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[3]/td').text
    custom_num = driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[4]/td').text
    # if name=="취소완료":
    #     return False
    print(id, name, address, phone_num, custom_num)
    add_customer(id, name, address, phone_num, custom_num)

def read_order_and_return_customer_count():
    driver.implicitly_wait(3)
    driver.switch_to.frame(driver.find_element_by_id("__naverpay")) #새로운 프레임으로 이동한다

    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')  #lxml 이 빠르므로 이걸로 쓰자
    content=[]
    list=soup.find_all(class_="ellipsis",columnname="PRODUCT_ORDER_ID",title=True)
    for item in list:
        content.append(item.text)
    customers_count=int(len(list))
    if customers_count==0:
        print("오늘은 하나도 팔리지않았군요")
        exit()

    #TODO : list 배열에서 title 값을 읽어오는 역할을 수행한다.
    for customer_num in range(0,customers_count):
        url = 'https://sell.smartstore.naver.com/o/orderDetail/productOrder/' + content[customer_num] + '/popup'  # 상세페이지 url
        open_info_pages(url)
    # for page in list:
    #
    #     open_info_pages(url)
    # for i in customer_count:
    return customers_count

def log_in(email,password):
    driver.implicitly_wait(3)
    driver.get('https://sell.smartstore.naver.com/#/login')
    driver.find_element_by_id('loginId').send_keys(email)
    driver.find_element_by_id('loginPassword').send_keys(password)
    driver.find_element_by_xpath('//*[@id="loginButton"]').click()
    driver.implicitly_wait(3)


#     in case when login doesnt work properly
def log_in_suc():
    if driver.current_url=='https://sell.smartstore.naver.com/#/login':
        return False
    return True
    #      TODO: 로그인 실패여부를 확인하는 함수를 만들어야한다 함수가 안정교함

def go_to_order_page():
    driver.implicitly_wait(3)
    if not log_in_suc():
        print("로그인 실패 수고링")
        exit()
    try:
        drop_down_menu = driver.find_element_by_xpath('//*[@id="seller-lnb"]/div/div[1]/ul/li[3]')
        drop_down_menu.click()
        driver.implicitly_wait(3)
        driver.find_element_by_css_selector(
        '#seller-lnb > div > div:nth-child(1) > ul > li.active > ul > li:nth-child(1) > a').click()
    except Exception as ex:
        print("실패",ex)
        exit()

def start_core_service(email,password):
    log_in(email,password)
    driver.implicitly_wait(3)
    go_to_order_page()
    customer_count = read_order_and_return_customer_count()
    file_name = os.path.dirname(os.path.realpath(__file__)) + get_excel_file_name()
    make_excel(file_name, customer_count)

if __name__=='__main__':
    # # TODO : 속도가 너무 느리다. 멀티프로세싱으로 좀 더 빠른 방법이 없을까?
    app = gui.QApplication(gui.sys.argv)
    ex = gui.MyApp()
    gui.sys.exit(app.exec_())

