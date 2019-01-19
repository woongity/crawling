import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver

driver = webdriver.Chrome('C:\python/chromedriver')

def wait():
    driver.implicitly_wait(3)
def long_wait():
    driver.implicitly_wait(8)

def ExcelExpress():
    wb= Workbook()
    ws1=wb.active
    ##ws2= wb.create_sheet(title="second_sheet")  새로운 시트 파일 하나 더 추가하기

    for row in range(1,10):
        for col in range (1,10):
            ws1.cell(row=row, column=col, value=int("{}{}".format(row,col)))

    wb.save("C:\study/test.xlsx")
##엑셀파일
def Login():
    wait()
    driver.get('https://www.hotdealzone.co.kr/')

    driver.find_element_by_name('mb_id').send_keys('ytw1122')
    driver.find_element_by_name('mb_password').send_keys('andrew3876')

    driver.find_element_by_xpath('//*[@id="ol_before"]/form/div/div[5]/button').click()

def ReadInfo():
    driver.find_element_by_xpath('//*[@id="main_wide"]/div[2]/div[1]/div[2]/span[1]/a').click()
    driver.find_element_by_xpath('//*[@id="content_wide"]/div/div[1]/ul/li[1]/a').click()

    long_wait()
    html=driver.page_source
    soup=BeautifulSoup(html,'html.parser')

    my_titles=driver.find_elements_by_class_name("auto_jak")
    #
    # my_titles=soup.select(
    #     '#nt_order_list_member > tr:nth-child(1) > td:nth-child(3) > table > tbody > tr > td.tracking > div:nth-child(1)'
    # )
    for item in my_titles:
        print(item.text)
ExcelExpress()
Login()
ReadInfo()
print("Crawling")
